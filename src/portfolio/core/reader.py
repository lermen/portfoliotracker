# This module is responsible for reading the Excel spreadsheet and turning its
# rows into typed Python objects (`Position` and `FixedIncomePosition`).
#
# It is the very first step in the data pipeline:
#   [Excel file] → reader.py → [list[Position]] → fetcher.py → engine.py → UI
#
# Design note: the public functions (`read_positions`, `read_fixed_income`) are
# async, but the actual file I/O happens in synchronous helpers (`_parse_excel`,
# `_parse_fixed_income`). This is intentional — openpyxl is a blocking library,
# so we run it in a background thread with `asyncio.to_thread` to avoid freezing
# the event loop. See `read_positions` for the full explanation.

# `asyncio` is Python's built-in library for writing concurrent code using the
# async/await syntax. "Concurrent" means multiple things can be in progress at
# the same time, even though Python typically runs one thread at a time.
import asyncio
from pathlib import Path

# `openpyxl` is a third-party library for reading and writing Excel (.xlsx) files.
# It lets us open a workbook, iterate over sheets, and read individual cell values.
import openpyxl

# `structlog` is a structured logging library. Instead of plain text log lines,
# it emits key=value pairs that are easy to filter and machine-parse in production.
import structlog

from portfolio.core.exceptions import ExcelParseError
from portfolio.core.models import FixedIncomePosition, Position

# Get a logger bound to this module. The `log` variable is used throughout the
# file to emit log events. Structured loggers let you attach arbitrary context:
#   log.info("event_name", key=value, another_key=value)
log = structlog.get_logger()


def _parse_excel(path: Path) -> list[Position]:
    """Open the first sheet of the Excel file and parse each row into a `Position`.

    Expected spreadsheet format (row 1 is the header, ignored):
      A: Ticker       — trading symbol, e.g. "AAPL" or "BTC-USD"
      B: Quantity     — number of shares/units (must be positive)
      C: Exchange     — optional display label, e.g. "NYSE", "B3"
      D: Category     — optional category, e.g. "Stocks", "Crypto", "FII"
      E: Avg Price    — optional average purchase price in native currency

    The leading underscore in `_parse_excel` is a Python convention signalling
    "private / internal use". Nothing stops external code from calling it, but
    the underscore communicates that it is an implementation detail not meant
    to be part of the public API.
    """
    try:
        # `read_only=True` streams the file row by row without loading it all into
        # memory at once — important for large files.
        # `data_only=True` returns the cached cell *values* instead of Excel formulas.
        # Without it, a cell containing `=A1*B1` would return the formula string,
        # not the computed number.
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        # Use the first sheet by position, not `wb.active`. `active` reflects
        # whichever tab was open when the file was last saved in Excel — if the
        # user saved while on the FixedIncome tab, `active` would point there
        # instead of the positions sheet.
        if not wb.worksheets:
            raise ExcelParseError("Workbook has no sheets")
        ws = wb.worksheets[0]

        positions: list[Position] = []

        # `iter_rows(min_row=2)` skips the header row (row 1) and yields each
        # subsequent row as a tuple of cell values.
        # `values_only=True` gives us the raw Python values rather than Cell objects.
        for row in ws.iter_rows(min_row=2, values_only=True):
            ticker, quantity = row[0], row[1]

            # Skip completely empty rows — openpyxl includes trailing blank rows
            # at the end of the sheet even if no data was entered there.
            if ticker is None or quantity is None:
                continue

            # `len(row) > 2` guards against rows that have fewer columns than expected.
            # The `and row[2] is not None` check prevents calling `.strip()` on None,
            # which would raise an `AttributeError`.
            exchange = (
                str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
            )
            category = (
                str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
            )
            avg_price_native = (
                float(row[4]) if len(row) > 4 and row[4] is not None else None
            )

            # Pydantic validates each field when `Position(...)` is called.
            # If `quantity` is <= 0, Pydantic raises a `ValidationError` automatically
            # because of the `Field(gt=0)` constraint in the model.
            positions.append(
                Position(
                    ticker=str(ticker).strip(),
                    quantity=float(quantity),
                    exchange=exchange,
                    category=category,
                    avg_price_native=avg_price_native,
                )
            )

        wb.close()
        log.info("excel_parsed", path=str(path), count=len(positions))
        return positions

    except ExcelParseError:
        # Re-raise our own exception type unchanged — we don't want to accidentally
        # wrap an `ExcelParseError` inside another `ExcelParseError` in the clause below.
        raise
    except Exception as exc:
        # Catch any other unexpected error and wrap it in our own type.
        # `raise ... from exc` preserves the original traceback as the "cause",
        # which is invaluable when debugging — it shows both errors in the output.
        raise ExcelParseError(f"Failed to parse {path}: {exc}") from exc


def _parse_fixed_income(path: Path) -> list[FixedIncomePosition]:
    """Parse the 'FixedIncome' sheet from the Excel file.

    The sheet is optional — if it doesn't exist, an empty list is returned so
    the rest of the app can run without modification.

    Expected columns (row 1 is the header):
      A: Name     — a descriptive label for the investment
      B: Amount   — the current BRL value of the position
    """
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        # `wb.sheetnames` is a list of sheet name strings in the workbook.
        # The `in` operator checks for membership — it reads like plain English:
        # "if 'FixedIncome' is not in the list of sheet names".
        if "FixedIncome" not in wb.sheetnames:
            wb.close()
            log.info("fixed_income_sheet_missing", path=str(path))
            return []   # empty list is always safe to iterate over, unlike None

        ws = wb["FixedIncome"]   # access a sheet by name using dictionary-style syntax
        items: list[FixedIncomePosition] = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            name, amount = row[0], row[1]
            if name is None or amount is None:
                continue
            items.append(
                FixedIncomePosition(
                    name=str(name).strip(),
                    amount_brl=float(amount),
                )
            )

        wb.close()
        log.info("fixed_income_parsed", path=str(path), count=len(items))
        return items

    except Exception as exc:
        raise ExcelParseError(f"Failed to parse FixedIncome sheet in {path}: {exc}") from exc


async def read_positions(path: Path) -> list[Position]:
    """Read variable-asset positions from the first sheet of the Excel file.

    This is an `async` function, meaning it must be called with `await`:
        positions = await read_positions(path)

    Why is it async when openpyxl is a blocking library?
    ----------------------------------------------------
    asyncio runs everything on a single thread. If you call a slow, blocking
    function directly (like reading a large file), the entire event loop — and
    every other async task, including UI updates — freezes until it finishes.

    `asyncio.to_thread(fn, *args)` solves this by running `fn` in a separate
    thread from a background thread pool. The event loop stays free to handle
    other coroutines while the file is being read in the background.
    """
    return await asyncio.to_thread(_parse_excel, path)


async def read_fixed_income(path: Path) -> list[FixedIncomePosition]:
    """Read fixed-income positions from the 'FixedIncome' sheet asynchronously.

    Uses the same `asyncio.to_thread` pattern as `read_positions` so the event
    loop is never blocked while the file is being read.
    """
    return await asyncio.to_thread(_parse_fixed_income, path)
